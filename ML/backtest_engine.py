import pandas as pd
import numpy as np
from . import financial_metrics

class BacktestEngine:
    def __init__(self, predictions, dates, daily_returns, target_horizon=21, feature_data=None, report_context=None):
        """
        predictions: Model scores
        dates: DateTime index
        daily_returns: Actual daily returns of SPY (aligned with dates)
        feature_data: Optional DataFrame with features (for regime analysis)
        report_context: Optional text/dict with context (Model description, Validation details)
        """
        # Align dataframes
        self.df = pd.DataFrame({
            'pred': predictions,
            'market_ret': daily_returns.values
        }, index=dates).dropna()
        self.target_horizon = target_horizon
        self.feature_data = feature_data
        self.report_context = report_context or {}

    def run_scenario(self, mode='daily', strategy='long_short', cost_bps=5.0):
        """
        mode: 'daily' (Overlapping) or 'monthly' (Rebalance every 21 days)
        strategy: 'long_short', 'long_only', 'big_move'
        """
        preds = self.df['pred']
        
        # 1. Signal Generation
        if strategy == 'long_short':
            signal = np.sign(preds)
        elif strategy == 'long_only':
            signal = pd.Series(np.where(preds > 0, 1.0, 0.0), index=preds.index)
        elif strategy == 'big_move':
            # Dynamic threshold or fixed
            thresh = 0.03 
            signal = np.zeros_like(preds)
            signal[preds > thresh] = 1.0
            signal[preds < -thresh] = -1.0
            signal = pd.Series(signal, index=preds.index)
        elif strategy == 'vol_targeting':
            # Volatility Targeting Strategy
            target_vol = 0.15  # 15% annualized vol target
            
            if self.feature_data is not None and 'GARCH_Forecast' in self.feature_data.columns:
                # Use GARCH forecast if available (annualized)
                # GARCH is usually variance, so sqrt it
                vol_est = np.sqrt(self.feature_data.loc[preds.index, 'GARCH_Forecast'])
            else:
                # Fallback: 21-day rolling vol
                vol_est = self.df['market_ret'].rolling(21).std() * np.sqrt(252)
            
            # Avoid division by zero
            vol_est = vol_est.replace(0, np.nan).ffill().fillna(0.15)
            
            # Size = Target / Estimate
            # Cap leverage at 2.0x to be realistic
            leverage = (target_vol / vol_est).clip(0, 2.0)
            
            # Direction from model prediction
            direction = np.sign(preds)
            
            # Ensure leverage is aligned (if vol_est came from rolling, it has index)
            if hasattr(leverage, 'values'):
                leverage = leverage.values
            
            signal = pd.Series(direction * leverage, index=preds.index)

        
        # 2. Execution Logic
        if mode == 'monthly':
            # Resample signal to every N days (snapshot)
            # This mimics taking a trade and holding it locked for N days
            sig_monthly = signal.iloc[::self.target_horizon]
            # Forward fill signal for holding period
            active_signal = sig_monthly.reindex(signal.index).ffill()
        else:
            # Daily Overlapping: Average of last N signals
            active_signal = signal.rolling(self.target_horizon).mean()
            
        # Shift 1 day (Trade next open) to prevent lookahead
        position = active_signal.shift(1).fillna(0)
        
        # 3. PnL Calculation
        gross_ret = position * self.df['market_ret']
        
        # Costs (Turnover * Cost)
        turnover = position.diff().abs().fillna(0)
        costs = turnover * (cost_bps / 10000)
        
        return gross_ret - costs

    def generate_boss_report_md(self):
        """Runs all scenarios and returns a Markdown table string."""
        scenarios = [
            ("L/S (Daily Overlap)", 'daily', 'long_short'),
            ("Long Only (Daily)", 'daily', 'long_only'),
            ("L/S (Monthly Rebal)", 'monthly', 'long_short'),
            ("Big Move (Daily)", 'daily', 'big_move'),
            ("Vol Target 15% (Daily)", 'daily', 'vol_targeting')
        ]

        
        # Run Benchmark (Buy & Hold)
        bench_ret = self.df['market_ret']
        bench_metrics = financial_metrics.generate_boss_metrics(bench_ret)
        
        # Prepare Data Rows
        rows = []
        # Add Benchmark
        rows.append(["**SPY Buy & Hold**"] + list(bench_metrics.values()))
        
        # Run Scenarios
        for name, mode, strat in scenarios:
            try:
                rets = self.run_scenario(mode, strat)
                m = financial_metrics.generate_boss_metrics(rets)
                rows.append([name] + list(m.values()))
            except Exception as e:
                print(f"Skipping scenario {name}: {e}")

        # Define Column Groups
        # Full keys: Total Return, CAGR, Volatility, Sharpe, Sortino, Calmar, Max DD, DD Duration, Win Rate, Profit Factor
        keys = list(bench_metrics.keys())
        
        # Table 1: Returns & Risk
        # Indices: 0, 1, 2, 3, 4, 5 (Total Return to Calmar)
        headers_1 = ["Strategy"] + keys[:6]
        
        # Table 2: Drawdown & Trade Stats
        # Indices: 6, 7, 8, 9 (Max DD to Profit Factor)
        headers_2 = ["Strategy"] + keys[6:]
        
        # Construct Markdown
        md = "\n## 💼 Boss Report: Trading Strategy Analysis\n"
        
        # --- Strategy Explanations ---
        md += "### 📘 Strategy Definitions\n"
        md += "To ensure full visibility on how these backtests operate, here are the detailed mechanics of each strategy:\n\n"
        
        md += "**1. L/S (Daily Overlap)**\n"
        md += "- **Concept**: Tells us how the model performs if we trade every single day based on the latest prediction.\n"
        md += "- **Mechanics**: Instead of putting 100% of capital into one trade, we enter 1/21th of the position each day. This smooths out volatility and noise. It effectively averages the signals over the last 21 days.\n"
        md += "- **Position**: Can be Long (+1), Short (-1), or Neutral. Net position fluctuates smoothly.\n\n"
        
        md += "**2. Long Only (Daily)**\n"
        md += "- **Concept**: Traditional 'Long Only' equity fund logic. No shorting allowed.\n"
        md += "- **Mechanics**: If the model predicts UP, we buy. If the model predicts DOWN, we go to Cash (0 position). Uses the same 'Daily Overlap' smoothing as above.\n"
        md += "- **Position**: Long (+1) or Cash (0).\n\n"
        
        md += "**3. L/S (Monthly Rebal)**\n"
        md += "- **Concept**: Standard monthly hedge fund rebalancing.\n"
        md += "- **Mechanics**: We look at the model's signal only once every 21 days. We ignore all daily fluctuations in between. We enter a trade and hold it 'locked' for the full month.\n"
        md += "- **Pros/Cons**: Lower transaction costs, but reacts slower to sudden market changes.\n\n"
        
        md += "**4. Big Move (Daily)**\n"
        md += "- **Concept**: 'Sniper' strategy. Only trade when the model is extremely confident.\n"
        md += "- **Mechanics**: We set a high threshold (e.g., 3%). If prediction > 3%, go 100% Long. If prediction < -3%, go 100% Short. Otherwise, sit in Cash.\n"
        md += "- **Position**: Binary bets: +1, -1, or 0.\n\n"

        md += "**5. Vol Target 15% (Daily)**\n"
        md += "- **Concept**: Risk parity approach. Trade larger when market is calm, smaller when volatile.\n"
        md += "- **Mechanics**: We target 15% annualized volatility. Leverage = 15% / Current Vol. Direction follows model.\n"
        md += "- **Position**: Variable leverage (Capped at 2x). E.g., if Vol is 30% -> 0.5x position. If Vol is 10% -> 1.5x position.\n\n"


        md += "**5. Vol Target 15% (Daily)**\n"
        md += "- **Concept**: Risk parity approach. Trade larger when market is calm, smaller when volatile.\n"
        md += "- **Mechanics**: We target 15% annualized volatility. Leverage = 15% / Current Vol. Direction follows model.\n"
        md += "- **Position**: Variable leverage (Capped at 2x). E.g., if Vol is 30% -> 0.5x position. If Vol is 10% -> 1.5x position.\n\n"

        
        md += "> **Note**: Simulation includes 5bps transaction costs per turnover (slippage + comms).\n\n"
        
        # --- Table 1 ---
        md += "### 📈 Returns & Risk Metrics\n"
        md += "| " + " | ".join(headers_1) + " |\n"
        md += "| " + " | ".join(["---"] * len(headers_1)) + " |\n"
        
        for row in rows:
            # Strategy Name + First 6 metrics
            table_row = [row[0]] + row[1:7]
            md += "| " + " | ".join(table_row) + " |\n"
            
        md += "\n"
        
        # --- Table 2 ---
        md += "### 📉 Drawdown & Trade Stats\n"
        md += "| " + " | ".join(headers_2) + " |\n"
        md += "| " + " | ".join(["---"] * len(headers_2)) + " |\n"
        
        for row in rows:
            # Strategy Name + Remaining metrics
            table_row = [row[0]] + row[7:]
            md += "| " + " | ".join(table_row) + " |\n"
            
        return md

    def _calculate_rv_ratio(self, returns, short_window=5, long_window=20):
        """Calculate RV_Ratio for regime detection."""
        rv_short = returns.rolling(window=short_window).std()
        rv_long = returns.rolling(window=long_window).std()
        return rv_short / rv_long

    def _get_regime_labels(self):
        """Determine volatility regime for each date."""
        if self.feature_data is not None and 'RV_Ratio' in self.feature_data.columns:
            # Use RV_Ratio from features if available
            aligned_rv = self.feature_data.loc[self.df.index, 'RV_Ratio']
        else:
            # Calculate from market returns
            aligned_rv = self._calculate_rv_ratio(self.df['market_ret'])
        
        # Use median as threshold (consistent with RegimeGatedRidge)
        threshold = aligned_rv.median()
        return (aligned_rv > threshold).map({True: 'High Vol', False: 'Low Vol'})

    def generate_plots(self, output_dir):
        """Generates and saves additional plots/visuals."""
        import os
        import matplotlib.pyplot as plt
        
        # Re-run scenarios to get returns
        scenarios = [
            ("L/S (Daily Overlap)", 'daily', 'long_short'),
            ("Long Only (Daily)", 'daily', 'long_only'),
            ("L/S (Monthly Rebal)", 'monthly', 'long_short'),
            ("Big Move (Daily)", 'daily', 'big_move'),
            ("Vol Target 15% (Daily)", 'daily', 'vol_targeting')
        ]
        
        strategy_returns = {
            "SPY Buy & Hold": self.df['market_ret']
        }
        for name, mode, strat in scenarios:
            try:
                strategy_returns[name] = self.run_scenario(mode, strat)
            except:
                pass
                
        # 1. Underwater Plot
        try:
            fig_uw = self.plot_underwater(strategy_returns)
            fig_uw.savefig(os.path.join(output_dir, "plot_underwater.png"))
            plt.close(fig_uw)
        except Exception as e:
            print(f"Failed to save underwater plot: {e}")

        # 2. Heatmap (for best strategy or just L/S)
        try:
            # Generate for Vol Target or L/S
            target_strat = "Vol Target 15% (Daily)"
            if target_strat not in strategy_returns:
                target_strat = "L/S (Daily Overlap)"
            
            if target_strat in strategy_returns:
                fig_hm = self.plot_monthly_heatmap(strategy_returns[target_strat])
                fig_hm.savefig(os.path.join(output_dir, "plot_monthly_heatmap.png"))
                plt.close(fig_hm)
        except Exception as e:
            print(f"Failed to save heatmap: {e}")

    def generate_boss_report_excel(self, save_path):
        """
        Generates comprehensive boss report with multiple sheets:
        1. Summary - Core metrics for all strategies
        2. Monthly Returns - Monthly breakdown
        3. Quarterly Returns - Quarterly breakdown
        4. Rolling Metrics - 6M and 12M rolling Sharpe
        5. Regime Analysis - Performance by volatility regime
        6. Trade Analysis - Trade-level statistics
        7. Period Sharpe - Annual/Quarterly Sharpe ratios
        """
        scenarios = [
            ("L/S (Daily Overlap)", 'daily', 'long_short'),
            ("Long Only (Daily)", 'daily', 'long_only'),
            ("L/S (Monthly Rebal)", 'monthly', 'long_short'),
            ("Big Move (Daily)", 'daily', 'big_move'),
            ("Vol Target 15% (Daily)", 'daily', 'vol_targeting')
        ]

        
        # Run Benchmark (Buy & Hold)
        bench_ret = self.df['market_ret']
        
        # Collect all strategy returns
        strategy_returns = {
            "SPY Buy & Hold": bench_ret
        }
        
        for name, mode, strat in scenarios:
            try:
                rets = self.run_scenario(mode, strat)
                strategy_returns[name] = rets
            except Exception as e:
                print(f"Skipping scenario {name}: {e}")
        
        # Generate all sheets
        sheets = {}
        
        # Sheet 0: Context / Read Me (The User Requested Overview)
        context_rows = []
        if self.report_context:
            for section, text in self.report_context.items():
                context_rows.append({"Section": section, "Details": text})
        
        # Add standard strategy definitions if not present
        if not any("Strategy" in str(k) for k in self.report_context.keys()):
             context_rows.append({"Section": "Strategy: L/S (Daily)", "Details": "Trades every day based on latest prediction (1/21 position size per day). Smooths volatility."})
             context_rows.append({"Section": "Strategy: Vol Target 15%", "Details": "Adjusts leverage dynamically to maintain 15% annualized volatility. Sizes down in high vol."})
        
        if context_rows:
            sheets['Report Context'] = pd.DataFrame(context_rows)

        # Sheet 1: Summary
        summary_data = []
        for name, rets in strategy_returns.items():
            metrics = financial_metrics.generate_boss_metrics_enhanced(rets, bench_ret)
            row = {"Strategy": name}
            row.update(metrics)
            summary_data.append(row)
        sheets['Summary'] = pd.DataFrame(summary_data)
        
        # Sheet 2: Monthly Returns
        monthly_data = []
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            monthly_rets = rets_series.resample('ME').apply(lambda x: np.prod(1 + x) - 1)
            for date, ret in monthly_rets.items():
                monthly_data.append({
                    "Strategy": name,
                    "Year": date.year,
                    "Month": date.month,
                    "Date": date.strftime('%Y-%m'),
                    "Return": ret
                })
        sheets['Monthly Returns'] = pd.DataFrame(monthly_data)
        
        # Sheet 3: Quarterly Returns
        quarterly_data = []
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            quarterly_rets = rets_series.resample('QE').apply(lambda x: np.prod(1 + x) - 1)
            for date, ret in quarterly_rets.items():
                quarter_num = (date.month-1)//3 + 1
                quarterly_data.append({
                    "Strategy": name,
                    "Year": date.year,
                    "Quarter": f"Q{quarter_num}",
                    "Date": f"{date.year}-Q{quarter_num}",
                    "Return": ret
                })
        sheets['Quarterly Returns'] = pd.DataFrame(quarterly_data)
        
        # Sheet 4: Rolling Metrics
        rolling_data = []
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            
            # 6-month rolling Sharpe (126 trading days)
            rolling_6m = rets_series.rolling(126)
            rolling_6m_ret = rolling_6m.mean() * 252
            rolling_6m_vol = rolling_6m.std() * np.sqrt(252)
            rolling_6m_sharpe = (rolling_6m_ret / rolling_6m_vol).replace([np.inf, -np.inf], np.nan)
            
            # 12-month rolling Sharpe (252 trading days)
            rolling_12m = rets_series.rolling(252)
            rolling_12m_ret = rolling_12m.mean() * 252
            rolling_12m_vol = rolling_12m.std() * np.sqrt(252)
            rolling_12m_sharpe = (rolling_12m_ret / rolling_12m_vol).replace([np.inf, -np.inf], np.nan)
            
            # Only include dates where we have valid rolling metrics (to reduce file size)
            valid_dates = rolling_6m_sharpe.dropna().index.union(rolling_12m_sharpe.dropna().index)
            for date in valid_dates:
                rolling_data.append({
                    "Strategy": name,
                    "Date": date,
                    "Rolling_6M_Sharpe": rolling_6m_sharpe.loc[date] if pd.notna(rolling_6m_sharpe.loc[date]) else None,
                    "Rolling_12M_Sharpe": rolling_12m_sharpe.loc[date] if pd.notna(rolling_12m_sharpe.loc[date]) else None,
                })
        sheets['Rolling Metrics'] = pd.DataFrame(rolling_data)
        
        # Sheet 5: Regime Analysis
        regime_labels = self._get_regime_labels()
        regime_data = []
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            
            # Group by regime
            for regime in ['Low Vol', 'High Vol']:
                regime_mask = regime_labels == regime
                regime_rets = rets_series[regime_mask]
                
                if len(regime_rets) > 0:
                    total_ret = np.prod(1 + regime_rets) - 1
                    ann_ret = regime_rets.mean() * 252
                    ann_vol = regime_rets.std() * np.sqrt(252)
                    sharpe = (ann_ret / ann_vol) if ann_vol > 0 else 0
                    
                    equity = (1 + regime_rets).cumprod()
                    max_dd, _ = financial_metrics.calculate_max_drawdown(equity)
                    
                    win_rate = (regime_rets > 0).sum() / len(regime_rets)
                    
                    regime_data.append({
                        "Strategy": name,
                        "Regime": regime,
                        "Total Return": total_ret,
                        "CAGR": ann_ret,
                        "Volatility": ann_vol,
                        "Sharpe Ratio": sharpe,
                        "Max Drawdown": max_dd,
                        "Win Rate": win_rate,
                        "Days": len(regime_rets)
                    })
        sheets['Regime Analysis'] = pd.DataFrame(regime_data)
        
        # Sheet 6: Trade Analysis (Strategy Level Stats) & Trade Log (Detailed)
        trade_data = []
        all_trades_log = []
        
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            
            # Identify trades (position changes)
            # This logic assumes daily rebalancing or effective position changes
            # For accurate trade logging, we need to reconstruct the position
            # This is an approximation based on returns != 0 for 'In Trade'
            # A better way for L/S strategies is to track sign changes
            
            positions = (rets_series != 0).astype(int)
            # For L/S, we want to track direction. 
            # If returns are + and market is +, we are Long. 
            # But simpler: use the predictions if available or infer from return correlation?
            # Since we only have 'rets' here, we infer 'Active' status.
            
            position_changes = positions.diff().abs()
            trade_starts = position_changes[position_changes > 0].index
            
            trades = []
            for i, start_date in enumerate(trade_starts):
                if i < len(trade_starts) - 1:
                    end_date = trade_starts[i + 1]
                else:
                    end_date = rets_series.index[-1]
                
                trade_rets = rets_series.loc[start_date:end_date]
                trade_return = np.prod(1 + trade_rets) - 1
                holding_days = len(trade_rets)
                
                # Heuristic for Direction: Correlation with Market
                market_period = bench_ret.loc[start_date:end_date]
                if len(market_period) > 1 and len(trade_rets) > 1:
                    # Sanitize
                    trade_rets_clean = np.nan_to_num(trade_rets, nan=0.0)
                    market_clean = np.nan_to_num(market_period, nan=0.0)
                    
                    if np.var(trade_rets_clean) > 1e-8 and np.var(market_clean) > 1e-8:
                        corr = np.corrcoef(trade_rets_clean, market_clean)[0, 1]
                        direction = "Long" if corr > 0 else "Short"
                    else:
                        direction = "Neutral"
                else:
                    direction = "Unknown"

                trade_entry = {
                    "Strategy": name,
                    "Entry Date": start_date,
                    "Exit Date": end_date,
                    "Direction": direction,
                    "Return": trade_return,
                    "Holding Days": holding_days
                }
                trades.append(trade_entry)
                all_trades_log.append(trade_entry)
            
            if trades:
                trades_df = pd.DataFrame(trades)
                largest_win = trades_df.loc[trades_df['Return'].idxmax()]
                largest_loss = trades_df.loc[trades_df['Return'].idxmin()]
                avg_holding = trades_df['Holding Days'].mean()
                
                trade_data.append({
                    "Strategy": name,
                    "Total Trades": len(trades),
                    "Largest Win": largest_win['Return'],
                    "Largest Win Start": largest_win['Entry Date'],
                    "Largest Loss": largest_loss['Return'],
                    "Largest Loss Start": largest_loss['Entry Date'],
                    "Avg Holding Days": avg_holding,
                    "Median Holding Days": trades_df['Holding Days'].median(),
                })
            else:
                trade_data.append({
                    "Strategy": name,
                    "Total Trades": 0,
                    "Largest Win": None,
                    "Largest Win Start": None,
                    "Largest Loss": None,
                    "Largest Loss Start": None,
                    "Avg Holding Days": None,
                    "Median Holding Days": None,
                })
        
        sheets['Trade Analysis'] = pd.DataFrame(trade_data)
        sheets['Trade Log'] = pd.DataFrame(all_trades_log) # New Sheet
        
        # Sheet 7: Period Sharpe
        period_sharpe_data = []
        for name, rets in strategy_returns.items():
            rets_series = pd.Series(rets, index=self.df.index)
            
            # Annual Sharpe (by calendar year)
            for year in rets_series.index.year.unique():
                year_rets = rets_series[rets_series.index.year == year]
                if len(year_rets) > 0:
                    ann_ret = year_rets.mean() * 252
                    ann_vol = year_rets.std() * np.sqrt(252)
                    sharpe = (ann_ret / ann_vol) if ann_vol > 0 else 0
                    
                    period_sharpe_data.append({
                        "Strategy": name,
                        "Period": str(year),
                        "Period Type": "Annual",
                        "Sharpe Ratio": sharpe,
                        "Return": np.prod(1 + year_rets) - 1,
                        "Volatility": ann_vol
                    })
            
            # Quarterly Sharpe
            quarterly_rets = rets_series.resample('QE').apply(lambda x: np.prod(1 + x) - 1)
            for date, q_ret in quarterly_rets.items():
                q_daily_rets = rets_series[rets_series.index.to_period('Q') == date.to_period('Q')]
                if len(q_daily_rets) > 0:
                    ann_ret = q_daily_rets.mean() * 252
                    ann_vol = q_daily_rets.std() * np.sqrt(252)
                    sharpe = (ann_ret / ann_vol) if ann_vol > 0 else 0
                    
                    period_sharpe_data.append({
                        "Strategy": name,
                        "Period": f"{date.year}-Q{(date.month-1)//3 + 1}",
                        "Period Type": "Quarterly",
                        "Sharpe Ratio": sharpe,
                        "Return": q_ret,
                        "Volatility": ann_vol
                    })
        
        sheets['Period Sharpe'] = pd.DataFrame(period_sharpe_data)
        
        # Sheet 8: Configuration (New)
        from . import config
        config_data = [
            {"Parameter": "Generated Date", "Value": str(pd.Timestamp.now())},
            {"Parameter": "Model Type", "Value": getattr(config, 'MODEL_TYPE', 'Unknown')},
            {"Parameter": "Frequency", "Value": getattr(config, 'DATA_FREQUENCY', 'daily')},
            {"Parameter": "Target Mode", "Value": getattr(config, 'TARGET_MODE', 'Unknown')},
            {"Parameter": "Transaction Cost (bps)", "Value": 5},
            {"Parameter": "Big Move Threshold", "Value": getattr(config, 'BIG_MOVE_THRESHOLD', 0.03)},
        ]
        sheets['Configuration'] = pd.DataFrame(config_data)

        # Generate Plots (Logic kept simpler here, returning dict of plot objects or saving internally?)
        # For this engine, we rely on the caller to handle plots typically, 
        # but let's add helper methods for the new requested visuals.
        
        # Save to Excel with Formatting
        try:
            # Try using xlsxwriter for advanced formatting
            with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
                # Write each sheet
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply formatting
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Define formats
                    wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top'})
                    header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
                    
                    # Apply header format
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    
                    # Special formatting for "Report Context"
                    if sheet_name == 'Report Context':
                        worksheet.set_column('A:A', 25, wrap_format)
                        worksheet.set_column('B:B', 100, wrap_format)
                    else:
                        # Auto-adjust column widths
                        for i, col in enumerate(df.columns):
                            try:
                                column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                            except:
                                column_len = 15
                            column_len = min(column_len, 50)
                            worksheet.set_column(i, i, column_len)
                            
            print(f"Boss report saved to: {save_path}")
            
        except ImportError:
            print("Note: 'xlsxwriter' not found. Using 'openpyxl' (standard) for Excel generation.")
            # Fallback to standard save with openpyxl formatting
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    for sheet_name, df in sheets.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Basic OpenPyXL Formatting
                    from openpyxl.utils import get_column_letter
                    for sheet_name in writer.sheets:
                        worksheet = writer.sheets[sheet_name]
                        if sheets[sheet_name].empty: continue
                        
                        for idx, col in enumerate(sheets[sheet_name].columns):
                            series = sheets[sheet_name][col]
                            try:
                                max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 2
                            except:
                                max_len = 15
                            max_len = min(max_len, 50)
                            col_letter = get_column_letter(idx + 1)
                            worksheet.column_dimensions[col_letter].width = max_len
                            
                print(f"Saved report (standard formatting) to: {save_path}")
                
            except Exception as e2:
                print(f"Critical error saving report: {e2}")

        except Exception as e:
            print(f"Error saving Excel report (xlsxwriter): {e}") 
            # Fallback for generic errors
            try:
                with pd.ExcelWriter(save_path) as writer:
                    for sheet_name, df in sheets.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Saved unformatted report to: {save_path}")
            except Exception as e3:
                 print(f"Critical error saving report: {e3}")
        
        return sheets

    def plot_underwater(self, returns_dict):
        """Generates an Underwater (Drawdown) Plot."""
        import matplotlib.pyplot as plt
        
        fig, ax = plt.subplots(figsize=(12, 6))
        
        for name, rets in returns_dict.items():
            equity = (1 + rets).cumprod()
            running_max = equity.cummax()
            drawdown = (equity / running_max) - 1
            
            # Highlight our strategies
            lw = 2 if "L/S" in name or "Vol" in name else 1
            alpha = 1.0 if "L/S" in name or "Vol" in name else 0.6
            
            ax.plot(drawdown.index, drawdown, label=name, linewidth=lw, alpha=alpha)
            
        ax.set_title("Underwater Plot (Drawdown over Time)")
        ax.set_ylabel("Drawdown %")
        ax.set_xlabel("Date")
        ax.grid(True, alpha=0.3)
        ax.legend(loc='lower left')
        ax.fill_between(drawdown.index, 0, -1, color='red', alpha=0.05) # Danger zone
        
        return fig

    def plot_monthly_heatmap(self, returns):
        """Generates a Monthly Returns Heatmap for a single strategy."""
        import seaborn as sns
        import matplotlib.pyplot as plt
        
        # Resample to monthly
        monthly_rets = returns.resample('ME').apply(lambda x: np.prod(1 + x) - 1)
        
        # Create Pivot Table: Year vs Month
        # Ensure we have month names
        df_heatmap = pd.DataFrame({'Return': monthly_rets})
        df_heatmap['Year'] = df_heatmap.index.year
        df_heatmap['Month'] = df_heatmap.index.month
        
        pivot = df_heatmap.pivot(index='Year', columns='Month', values='Return')
        
        fig, ax = plt.subplots(figsize=(10, len(pivot)*0.5 + 2))
        sns.heatmap(pivot, annot=True, fmt='.1%', cmap='RdYlGn', center=0, cbar=False, ax=ax)
        ax.set_title("Monthly Returns Heatmap")
        
        return fig

        # We need to pass config settings in, or grab them from global config if possible.
        # Ideally, we pass them in __init__, but for now we'll hardcode what we know from the class state + defaults.
        from . import config
        config_data = [
            {"Parameter": "Generated Date", "Value": str(pd.Timestamp.now())},
            {"Parameter": "Target Horizon", "Value": f"{self.target_horizon} Days"},
            {"Parameter": "Transaction Cost", "Value": "5 bps"},
            {"Parameter": "Execution Frequency", "Value": str(getattr(config, 'EXECUTION_FREQUENCY', 'daily'))},
            {"Parameter": "Model Type", "Value": str(getattr(config, 'MODEL_TYPE', 'Unknown'))},
            {"Parameter": "Training Window", "Value": str(getattr(config, 'TRAIN_WINDOW_YEARS', 'Unknown')) + " Years"},
            {"Parameter": "Big Move Threshold", "Value": str(getattr(config, 'BIG_MOVE_THRESHOLD', 0.03))},
            {"Parameter": "Regime Logic", "Value": "RV_Ratio (Median Split)"},
            {"Parameter": "Data Embargo", "Value": str(getattr(config, 'EMBARGO_ROWS', 0)) + " Days"},
        ]
        sheets['Configuration'] = pd.DataFrame(config_data)
        
        # Save to Excel
        excel_path = save_path if save_path.endswith('.xlsx') else save_path + '.xlsx'
        csv_path = save_path.replace('.xlsx', '.csv') if save_path.endswith('.xlsx') else save_path + '.csv'
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatting
                    worksheet = writer.sheets[sheet_name]
                    
                    # Auto-adjust column widths
                    from openpyxl.utils import get_column_letter
                    for idx, col in enumerate(df.columns, start=1):
                        max_length = max(
                            df[col].astype(str).map(len).max(),
                            len(str(col))
                        )
                        col_letter = get_column_letter(idx)
                        col_letter = get_column_letter(idx)
                        # Specific width for Context sheet details
                        if sheet_name == 'Report Context' and col == 'Details':
                            worksheet.column_dimensions[col_letter].width = 100
                            # Wrap text for details
                            from openpyxl.styles import Alignment
                            for row in range(2, len(df) + 2):
                                cell = worksheet.cell(row=row, column=idx)
                                cell.alignment = Alignment(wrap_text=True)
                        else:
                            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
                    
                    # Format header
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format percentage columns
                    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00
                    pct_keywords = ['return', 'cagr', 'volatility', 'dd', 'win rate', 'sharpe']
                    for col_idx, col_name in enumerate(df.columns):
                        if any(kw in str(col_name).lower() for kw in pct_keywords):
                            for row in range(2, len(df) + 2):
                                cell = worksheet.cell(row=row, column=col_idx + 1)
                                if isinstance(cell.value, (int, float)):
                                    # Sharpe is not a percent
                                    if 'sharpe' in str(col_name).lower():
                                        cell.number_format = FORMAT_NUMBER_00
                                    elif abs(cell.value) < 10: # Reasonable check for percent
                                        cell.number_format = FORMAT_PERCENTAGE_00
            
            print(f"✅ Boss Report Excel saved to: {excel_path}")
            
            # Also save Summary as CSV for quick viewing
            sheets['Summary'].to_csv(csv_path, index=False)
            print(f"✅ Boss Report CSV (Summary) saved to: {csv_path}")
            
        except ImportError:
            print("⚠️  openpyxl not installed. Saving as CSV only. Install with: pip install openpyxl")
            sheets['Summary'].to_csv(csv_path, index=False)
            print(f"✅ Boss Report CSV saved to: {csv_path}")
        
        return sheets
